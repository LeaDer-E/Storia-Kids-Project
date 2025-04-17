import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from datetime import datetime

# Find the Excel file named 'Orders.xlsx'
file_path = None
for filename in os.listdir('.'):
    if filename == 'Orders.xlsx':
        file_path = filename
        break

if not file_path:
    raise FileNotFoundError("No file named 'Orders.xlsx' found.")

# Load the Excel file
df = pd.read_excel(file_path)

# Filter the data
filtered_df = df[
    (df['Operation Confirmation: '] == 'Confirmed') &
    (df['Order Status: '] == 'Sent to Delivery') &
    (df['Delivery Status: '].isin([
        'Edit Completed', 'No Response', 'Pending Action', 
        'Pending Another Story', 'Pending Customer Confirmation', 
        'Pending Edit', 'Pending Phone Number', 'Customer Cancelled'
    ]))
]

# Select and reorder the columns
final_df = filtered_df[['OrderID: ', 'Phone: ', 'Country: ', 'Package: ', 'Department: ', 
                        'Delivery Status: ', 'Confirmation Date: ', 'ETA: ',
                        'Payment Method: ' ,'Payment Status: ',  'Left to Pay: ']].copy()

# Remove the time from the date columns (Confirmation Date and ETA)
final_df['Confirmation Date: '] = pd.to_datetime(final_df['Confirmation Date: ']).dt.date
final_df['ETA: '] = pd.to_datetime(final_df['ETA: ']).dt.date

# Create a new Excel file name with date and time (hour, minute, second)
today_date = datetime.now().strftime("%m-%d %H-%M-%S")
output_file = f"FULL Sheet {today_date}.xlsx"

# Save the filtered DataFrame to an Excel file using openpyxl
wb = Workbook()
ws = wb.active
ws.title = "Filtered Orders"

# Write DataFrame to Excel (rows)
for r in dataframe_to_rows(final_df, index=False, header=True):
    ws.append(r)

# Set column widths
ws.column_dimensions['A'].width = 25.00
ws.column_dimensions['B'].width = 45.00
ws.column_dimensions['C'].width = 20.00
ws.column_dimensions['D'].width = 30.00
ws.column_dimensions['E'].width = 35.00
ws.column_dimensions['F'].width = 80.00
ws.column_dimensions['G'].width = 50.00
ws.column_dimensions['H'].width = 50.00
ws.column_dimensions['I'].width = 30.00
ws.column_dimensions['J'].width = 30.00
ws.column_dimensions['K'].width = 30.00


# Freeze the first row and first column
ws.freeze_panes = 'B2'

# Apply borders to all cells
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Set cell styles for all cells (black background, white font, centered text, 30 pt font)
black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
white_font = Font(color="FFFFFF", size=30)
alignment_center = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
    for cell in row:
        cell.fill = black_fill
        cell.font = white_font
        cell.border = thin_border
        cell.alignment = alignment_center

# Apply special styles for the first row (header)
header_font = Font(bold=True, color="FFA500", size=30)  # Light orange color for the header row
ws.row_dimensions[1].height = 75  # Set first row height to 100

for cell in ws[1:1]:
    cell.font = header_font

# Apply special styles for Column C1 (header "Package"), but keep background black
ws['C1'].fill = black_fill  # Black background
ws['C1'].font = header_font  # Light orange font (same as first row)

# Apply special styles for the "Package" column (Column C - except header)
gold_font = Font(color="FFD700", size=30)  # Gold color for Package column
for cell in ws['D'][1:]:
    cell.font = gold_font

# Save the file
wb.save(output_file)
print(f"Filtered data saved to {output_file}")

