import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from datetime import datetime

def process_orders(input_file, output_file):
    # Read the input Excel file
    df = pd.read_excel(input_file)

    # Ensure that column names are stripped of extra spaces
    df.columns = df.columns.str.strip()

    # Extract columns
    order_id_col = 'OrderID:'
    chat_ref_col = 'Chat Reference:'
    delivery_status_col = 'Delivery Status:'
    phone_col = 'Phone:'
    department_col = 'Department:'
    package_col = 'Package:'
    country_col = 'Country:'

    # Filter orders with 'Pending Edit'
    pending_another_story = df[df[delivery_status_col] == 'Pending Edit']

    # Dictionary to hold the merged data
    output_data = []

    # Process each unique Chat Reference for 'Pending Edit'
    for chat_ref in pending_another_story[chat_ref_col].unique():
        # Filter orders with the same Chat Reference
        group = df[df[chat_ref_col] == chat_ref]

        # Get 'Pending Edit' Orders
        pending_orders = group[group[delivery_status_col] == 'Pending Edit']
        related_orders = group[group[delivery_status_col] != 'Pending Edit']

        # Prepare merged Order IDs
        if not pending_orders.empty:
            merged_ids = ', '.join(map(str, pending_orders[order_id_col]))
            output_data.append([merged_ids])

        # Add related orders to output
        for _, row in related_orders.iterrows():
            output_data.append([
                row[order_id_col],
                row[delivery_status_col],
                row[phone_col],
                row[department_col],
                row[package_col],
                row[country_col]
            ])
            
        # Add '-----------------------------------------------------------------------------------------------------------------------' if there are pending orders
        if not pending_orders.empty and not related_orders.empty:
            output_data.append(['-----------------------------------------------------------------------------------------------------------------------'])

    # Create DataFrame for output
    output_df = pd.DataFrame(output_data, columns=['Order IDs', 'Delivery Status', 'Phone', 'Department', 'Package', "Country"])

    # Save the output to an Excel file
    output_df.to_excel(output_file, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    light_yellow_fill = PatternFill(start_color='FFFFC5', end_color='FFFFC5', fill_type='solid') 
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    centered_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    no_border = Border()  # Define no border
    bold_font = Font(bold=True)  # Define bold font

    # Set column widths
    ws.column_dimensions['A'].width = 10.00
    ws.column_dimensions['B'].width = 30.50
    ws.column_dimensions['C'].width = 18.00
    ws.column_dimensions['D'].width = 12.50
    ws.column_dimensions['E'].width = 14.00

    # Apply formatting for cells
    for row in ws.iter_rows(min_row=2, max_col=6):
        for cell in row:
            cell.alignment = centered_alignment
            cell.border = thin_border

    # Apply bold font to Order ID cells
    for row in ws.iter_rows(min_row=2, max_col=6):
        for cell in row:
            if cell.column == 1:  # Apply bold font to Order ID column (Column A)
                cell.font = bold_font

    # Apply formatting for merged cells
    row_index = 2
    while row_index <= ws.max_row:
        if ws.cell(row=row_index, column=1).value == '-----------------------------------------------------------------------------------------------------------------------':
            # Merge cells from A to E for rows with '-----------------------------------------------------------------------------------------------------------------------'
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=6)
            row_index += 1
            continue

        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=6)
        ws.cell(row=row_index, column=1).fill = light_yellow_fill
        row_index += 1

        while row_index <= ws.max_row and ws.cell(row=row_index, column=1).value != '-----------------------------------------------------------------------------------------------------------------------':
            row_index += 1

    # Apply yellow fill to the frozen row
    for col in range(1, 7):  # Columns A to E
        ws.cell(row=1, column=col).fill = yellow_fill
        
    # Freeze the first row
    ws.freeze_panes = 'A2'

    # Save the workbook with formatting
    wb.save(output_file)

# Generate a timestamp for the output file
now = datetime.now()
timestamp = now.strftime("%m.%d-%H.%M.%S")
output_file = f'PE {timestamp}.xlsx'

# Process the orders
process_orders('Orders.xlsx', output_file)

