import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from datetime import datetime

def process_orders(input_file, output_file):
    # Read the input Excel file
    df = pd.read_excel(input_file)

    # Ensure that column names are stripped of extra spaces
    df.columns = df.columns.str.strip()

    # Filter orders
    df_filtered = df[
        (df['Operation Confirmation:'] == 'Confirmed') &
        (df['Order Status:'] == 'Sent to Delivery') &
        (df['Delivery Status:'] == 'No Response')
    ]

    # Group by Phone, and collect related order details
    unique_phones = df_filtered['Phone:'].unique()
    output_data = []

    for phone in unique_phones:
        # Filter rows with the same phone
        phone_group = df_filtered[df_filtered['Phone:'] == phone]

        # Add the phone number in light yellow merged cells
        output_data.append([phone, '', '', '', '', ''])  # Placeholder for phone row
        
        # Add the related orders for this phone number
        for _, row in phone_group.iterrows():
            output_data.append([
                row['OrderID:'], row['Department:'], row['Package:'], 
                row['Country:'], row['Confirmation Date:'], row['ETA:']
            ])

    # Create DataFrame for output
    output_df = pd.DataFrame(output_data, columns=['Order IDs', 'Department', 'Package', 'Country', 'Confirmation Date', 'ETA'])

    # Save to Excel
    output_df.to_excel(output_file, index=False)

    # Load the workbook and apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    light_yellow_fill = PatternFill(start_color='FFFFC5', end_color='FFFFC5', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    centered_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    bold_font = Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 10.00
    ws.column_dimensions['B'].width = 15.00
    ws.column_dimensions['C'].width = 15.00
    ws.column_dimensions['D'].width = 8.00
    ws.column_dimensions['E'].width = 20.00
    ws.column_dimensions['F'].width = 20.00

    # Apply formatting to header
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.alignment = centered_alignment
        cell.border = thin_border
        cell.font = bold_font
        
        # Set header height to 2x the default
        ws.row_dimensions[1].height = ws.row_dimensions[1].height * 2 if ws.row_dimensions[1].height else 1.0

    # Apply formatting for merged phone number cells
    row_index = 2
    while row_index <= ws.max_row:
        phone_value = ws.cell(row=row_index, column=1).value
        if phone_value and not ws.cell(row=row_index, column=2).value:  # If phone number row
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=6)
            ws.cell(row=row_index, column=1).fill = light_yellow_fill
            ws.cell(row=row_index, column=1).alignment = centered_alignment
            
            # Set row height to 2x the default
            ws.row_dimensions[row_index].height = ws.row_dimensions[row_index].height * 1 if ws.row_dimensions[row_index].height else 15.0
            
            # Apply borders to the merged cell
            for col in range(1, 7):  # Columns A to F
                ws.cell(row=row_index, column=col).border = thin_border
            
            row_index += 1
        else:
            # Apply border and alignment to order data rows
            for cell in ws[row_index]:
                cell.alignment = centered_alignment
                cell.border = thin_border
            row_index += 1

    # Apply borders to all header cells
    for cell in ws[1]:
        cell.border = thin_border

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Save the workbook with formatting
    wb.save(output_file)

# Generate a timestamp for the output file
now = datetime.now()
timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
output_file = f'NR_Customers_{timestamp}.xlsx'

# Process the orders
process_orders('Orders.xlsx', output_file)

